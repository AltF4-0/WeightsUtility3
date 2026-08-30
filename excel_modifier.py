#!/usr/bin/env python3
def modify_excel(path=None):

    import datetime
    import math
    from pathlib import Path
    from typing import cast

    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.cell.cell import MergedCell
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.page import PageMargins

    rows_per_page = 49

    # Offset for the pallet total rows
    BOTTOM_OFFSET = 3

    # Columns widths A   B   C   D   E   F   G
    column_widths = [12, 11, 9, 13, 12, 13, 22]

    # Paths
    today = datetime.datetime.now(datetime.timezone.utc).astimezone().strftime("%Y_%m_%d")
    save_path = Path.home() / "Desktop" / f"WeightSheet_{today}.xlsx"

    # Checks for alternate input path from GUI script, if none is found then it defaults to the downloads folder
    if path:
            latest_xls = Path(path)
            if not latest_xls.is_file():
                raise FileNotFoundError(f"Provided file does not exist: {latest_xls}")
    else:
        downloads = Path.home() / "Downloads"
        xls_files = list(downloads.glob("*.xls*"))
        if not xls_files:
            raise FileNotFoundError("No .xls files found in location!")
        latest_xls = max(xls_files, key=lambda f: f.stat().st_mtime)

    # Load spreadsheet into pandas
    weight_sheet = pd.read_excel(
        latest_xls,
        skiprows=1,
        header=7,
        index_col=None,
        na_values=["NA"]
    )

    # Clean up unnecessary data inside the dataframe
    permitted_entries = ["Kikuyu", "Kings Pride", "Santa Anna", "Kenda"]

    weight_sheet.replace("Kikuyu - Fine Leaf", "Kikuyu", inplace=True)
    weight_sheet = (
        weight_sheet
        .drop(columns=["PAD LOC", "Est. KG's", "TRANSPORT"], errors="ignore")
        .dropna(how="all").reset_index(drop=True)
        .loc[lambda df: df.iloc[:, 1].astype(str).str.strip().isin(permitted_entries)]
        .reset_index(drop=True)
    )

    # Pallet logic
    MAXI_PALLET = 20
    SPECIAL_PALLET = 40
    NORMAL_PALLET = 50

    result_rows = []

    for _, row in weight_sheet.iterrows():
        qty = int(row["QTY"])
        pallets = []

        if "Maxi" in str(row.get("ROLL TYPE", "")):
            num_pallets = math.ceil(qty / MAXI_PALLET)
            pallets = [MAXI_PALLET] * num_pallets
            pallets[-1] = qty - MAXI_PALLET * (num_pallets - 1)
        else:
            while qty > 0:
                if 50 < qty < 60:
                    pallets.extend([SPECIAL_PALLET, qty - SPECIAL_PALLET])
                    break
                elif qty >= NORMAL_PALLET:
                    pallets.append(NORMAL_PALLET)
                    qty -= NORMAL_PALLET
                else:
                    pallets.append(qty)
                    break

        row["PAL QTY"] = len(pallets)
        row["PALLET QTY"] = pallets[0]
        result_rows.append(row.copy())

        for p in pallets[1:]:
            blank = pd.Series(dtype=object)
            blank["PAL QTY"] = len(pallets)
            blank["PALLET QTY"] = p
            result_rows.append(blank)

    # Output Dataframe
    df = pd.DataFrame(result_rows).reset_index(drop=True)
    df.insert(loc=7, column="G", value="")
    cols = df.columns.tolist()
    cols.remove("PALLET QTY")
    qty_index = cols.index("QTY")
    cols.insert(qty_index + 1, "PALLET QTY")
    df = df[cols]
    if "PAL QTY" in df.columns:
        df.drop(columns=["PAL QTY"], inplace=True)

    # Creates a seperate copy of the DataFrame in order to count the pallet totals (Gets deleted)
    turf_pallet = cast(pd.DataFrame, df[["TURF", "PALLET TYPE"]])
    df_count = turf_pallet.ffill()
    by_variety = df_count.groupby(["TURF", "PALLET TYPE"]).size()

    def vc(turf, pallet_type):
        try:
            return by_variety[(turf, pallet_type)]
        except KeyError:
            return 0

    # Insert pallet totals
    totals_blocks = [
        # Header
        {"INV#": "TOTALS", "TURF": "SLIDE", "QTY": "PLAIN", "PALLET QTY": "PINK"},
        # Kikuyu
        {"INV#": "Kikuyu",      "TURF": vc("Kikuyu", "SLIDE"),      "QTY": vc("Kikuyu", "PLAIN"),      "PALLET QTY": vc("Kikuyu", "PINK")},
        # Kings Pride
        {"INV#": "Kings Pride", "TURF": vc("Kings Pride", "SLIDE"),  "QTY": vc("Kings Pride", "PLAIN"), "PALLET QTY": vc("Kings Pride", "PINK")},
    ]

    totals_df = pd.DataFrame(totals_blocks)

    # Logic for how many rows of cells are drawn with borders
    pages = math.ceil(len(df) / rows_per_page) or 1
    MAX_FORMAT_ROWS = rows_per_page * pages
    insert_at = MAX_FORMAT_ROWS - BOTTOM_OFFSET - 1

    if len(df) > insert_at:
        pages += 1
        MAX_FORMAT_ROWS = rows_per_page * pages
        insert_at = MAX_FORMAT_ROWS - BOTTOM_OFFSET - 1

    if insert_at > len(df):
        padding = pd.DataFrame(
            [[None] * len(df.columns)] * (insert_at - len(df)),
            columns=df.columns
        )
        df = pd.concat([df, padding], ignore_index=True)

    df = pd.concat([
        df.iloc[:insert_at],
        totals_df,
        df.iloc[insert_at + BOTTOM_OFFSET:]
    ], ignore_index=True)

    # Export modified spreadsheet to save path
    df.to_excel(save_path, index=False)

    # OpenPyXl formatting
    wb = load_workbook(save_path)
    ws = wb.active
    assert ws is not None, "Workbook has no active worksheet"

    center_align = Alignment(horizontal="center", vertical="center")

    grey_fill = PatternFill(
            start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
        )
    dark_grey_fill = PatternFill(
            start_color="9A9A9A", end_color="9A9A9A", fill_type="solid"
        )

    thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )

    header_fill = PatternFill(
            start_color="222222", end_color="222222", fill_type="solid"
        )

    for row in ws.iter_rows(
        min_row=1,
        max_row=MAX_FORMAT_ROWS,
        min_col=1,
        max_col=ws.max_column
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            if cell.value is not None:
                cell.fill = grey_fill
            if cell.value == "Kings Pride":
                cell.fill = dark_grey_fill

    # Bold white text on dark grey fill for header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")

    # Column G ("WEIGHTS") is filled with an invisable charater so LibreOffice actually prints the borders of the blank cells
    g_col_index = 7
    for row_idx in range(2, MAX_FORMAT_ROWS + 1):
        cell = ws.cell(row=row_idx, column=g_col_index)
        if not isinstance(cell, MergedCell) and cell.value is None:
            cell.value = " "

    # Adjust column widths and affix date to "WEIGHTS" column header
    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws["G1"].value = f"WEIGHTS {today}"

    # Set page margins
    ws.page_margins = PageMargins(
        left=0.5,
        right=0.5,
        top=0.8,
        bottom=0.5,
        header=0.5,
        footer=0.5,
    )

    # Save workbook
    wb.save(save_path)
